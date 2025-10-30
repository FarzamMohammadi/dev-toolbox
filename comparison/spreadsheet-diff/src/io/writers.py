"""
Output writers for comparison results.
Supports CSV, Excel, and HTML formats.
"""

from pathlib import Path
from datetime import datetime
from typing import Optional, List
from dataclasses import dataclass
import polars as pl
from rich.console import Console

from ..config.settings import ComparisonSettings, FileFormat


console = Console()


@dataclass
class SearchPanesConfig:
    """
    Configuration for DataTables SearchPanes extension.
    Enables interactive column filtering in HTML reports.
    """

    enabled: bool
    columns: List[int]  # Column indices to show filters for
    cascading: bool     # Enable cascading filters
    view_total: bool    # Show total counts
    threshold: float    # Auto-collapse threshold

    @classmethod
    def from_settings(
        cls,
        settings: ComparisonSettings,
        column_names: List[str],
        key_column_name: str
    ) -> "SearchPanesConfig":
        """
        Create SearchPanesConfig from ComparisonSettings.

        Args:
            settings: Comparison settings
            column_names: List of all column names in the output
            key_column_name: Name of the key column

        Returns:
            SearchPanesConfig instance
        """
        if not settings.enable_search_panes:
            return cls(
                enabled=False,
                columns=[],
                cascading=False,
                view_total=False,
                threshold=1.0
            )

        # Determine which columns to show filters for
        filter_column_names = settings.get_search_panes_filter_columns()

        if filter_column_names is None:
            # Auto-detect: key column, "field", and "type" columns
            # Skip "source_value" and "comparison_value" (usually too many unique values)
            filter_column_names = []
            if key_column_name in column_names:
                filter_column_names.append(key_column_name)
            if "field" in column_names:
                filter_column_names.append("field")
            if "type" in column_names:
                filter_column_names.append("type")

        # Convert column names to indices
        column_indices = []
        for col_name in filter_column_names:
            if col_name in column_names:
                column_indices.append(column_names.index(col_name))

        return cls(
            enabled=True,
            columns=column_indices,
            cascading=settings.search_panes_cascading,
            view_total=settings.search_panes_view_total,
            threshold=settings.search_panes_threshold
        )

    def to_js_config(self) -> str:
        """
        Generate JavaScript configuration object for DataTables SearchPanes.

        Returns:
            JavaScript object string
        """
        if not self.enabled or not self.columns:
            return ""

        # Format: columns: [0, 1, 4]
        columns_js = f"[{', '.join(map(str, self.columns))}]"

        config = f"""
        searchPanes: {{
            cascadePanes: {'true' if self.cascading else 'false'},
            viewTotal: {'true' if self.view_total else 'false'},
            threshold: {self.threshold},
            columns: {columns_js},
            layout: 'columns-3',
            clear: true,
            initCollapsed: false,
            dtOpts: {{
                searching: true,
                paging: true,
                pagingType: 'numbers',
                pageLength: 10,
                dom: 'tp'
            }}
        }},
        columnDefs: [{{
            searchPanes: {{
                show: true,
                orthogonal: 'filter'
            }},
            targets: {columns_js}
        }}],
        language: {{
            searchPanes: {{
                title: {{
                    _: 'Filters Active - %d',
                    0: 'No Filters Active',
                    1: 'Filter Active - 1'
                }},
                clearMessage: 'Clear All',
                collapse: {{
                    0: 'Search Filters',
                    _: 'Search Filters (%d)'
                }}
            }}
        }}"""

        return config

    def get_cdn_links(self) -> dict:
        """
        Get CDN links for SearchPanes extension.

        Returns:
            Dictionary with 'css' and 'js' keys containing CDN URLs
        """
        if not self.enabled:
            return {"css": [], "js": []}

        return {
            "css": [
                "https://cdn.datatables.net/select/1.7.0/css/select.dataTables.min.css",
                "https://cdn.datatables.net/searchpanes/2.2.0/css/searchPanes.dataTables.min.css"
            ],
            "js": [
                "https://cdn.datatables.net/select/1.7.0/js/dataTables.select.min.js",
                "https://cdn.datatables.net/searchpanes/2.2.0/js/dataTables.searchPanes.min.js"
            ]
        }

    def get_dom_layout(self) -> str:
        """
        Get DataTables DOM layout string.

        Returns:
            DOM layout string ('Pfrtip' with SearchPanes, 'frtip' without)
        """
        if self.enabled and self.columns:
            return "Pfrtip"  # P = SearchPanes, f = filter, r = processing, t = table, i = info, p = pagination
        return "frtip"  # Standard layout without SearchPanes


class ResultWriter:
    """
    Write comparison results to various output formats.
    """

    def __init__(self, settings: ComparisonSettings):
        """
        Initialize result writer.

        Args:
            settings: Comparison settings
        """
        self.settings = settings
        self.output_dir = settings.output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def write_differences(
        self,
        differences: pl.DataFrame,
        source_file: str,
        comparison_file: str,
        summary_stats: Optional[dict] = None
    ) -> List[Path]:
        """
        Write comparison differences to output files.

        Args:
            differences: DataFrame containing differences
            source_file: Name of source file
            comparison_file: Name of comparison file
            summary_stats: Optional summary statistics

        Returns:
            List of output file paths
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_files = []

        if self.settings.output_format in ["csv", "both"]:
            csv_file = self._write_csv(differences, timestamp)
            output_files.append(csv_file)

        if self.settings.output_format in ["excel", "both"]:
            excel_file = self._write_excel(differences, timestamp)
            output_files.append(excel_file)

        if self.settings.generate_html_report:
            html_file = self._write_html_report(
                differences,
                source_file,
                comparison_file,
                timestamp,
                summary_stats
            )
            output_files.append(html_file)

        return output_files

    def _write_csv(self, df: pl.DataFrame, timestamp: str) -> Path:
        """Write differences to CSV file."""
        output_file = self.output_dir / f"differences_{timestamp}.csv"

        df.write_csv(output_file)

        console.print(f"[green]CSV saved:[/green] {output_file}")
        return output_file

    def _write_excel(self, df: pl.DataFrame, timestamp: str) -> Path:
        """Write differences to Excel file with formatting."""
        output_file = self.output_dir / f"differences_{timestamp}.xlsx"

        # Check row limit
        if len(df) > FileFormat.EXCEL_MAX_ROWS:
            console.print(
                f"[yellow]Warning: {len(df):,} differences exceed Excel's "
                f"{FileFormat.EXCEL_MAX_ROWS:,} row limit. "
                f"Truncating to fit Excel format.[/yellow]"
            )
            df = df.head(FileFormat.EXCEL_MAX_ROWS)

        # Write using xlsxwriter for better formatting control
        with pl.Config() as cfg:
            df.write_excel(
                output_file,
                worksheet="Differences",
                autofit=True
            )

        # Apply conditional formatting if possible
        try:
            self._apply_excel_formatting(output_file, df)
        except Exception as e:
            console.print(f"[yellow]Note: Could not apply Excel formatting: {e}[/yellow]")

        console.print(f"[green]Excel saved:[/green] {output_file}")
        return output_file

    def _apply_excel_formatting(self, filepath: Path, df: pl.DataFrame):
        """Apply conditional formatting to Excel file."""
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill

        wb = load_workbook(filepath)
        ws = wb.active

        # Define fill colors
        source_fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
        comparison_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')

        # Find columns (assuming standard naming)
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

        source_col_idx = None
        comparison_col_idx = None

        for idx, header in enumerate(header_row, 1):
            if header and 'SOURCE' in str(header).upper():
                source_col_idx = idx
            elif header and ('COMPARISON' in str(header).upper() or 'NEW' in str(header).upper()):
                comparison_col_idx = idx

        # Apply formatting
        for row_idx in range(2, len(df) + 2):  # Skip header
            if source_col_idx:
                ws.cell(row=row_idx, column=source_col_idx).fill = source_fill
            if comparison_col_idx:
                ws.cell(row=row_idx, column=comparison_col_idx).fill = comparison_fill

        wb.save(filepath)

    def _write_html_report(
        self,
        df: pl.DataFrame,
        source_file: str,
        comparison_file: str,
        timestamp: str,
        summary_stats: Optional[dict] = None
    ) -> Path:
        """Generate interactive HTML report with DataTables and SearchPanes filtering."""
        output_file = self.output_dir / f"differences_report_{timestamp}.html"

        # Limit rows for HTML (performance)
        max_html_rows = 50000
        is_truncated = False
        if len(df) > max_html_rows:
            console.print(
                f"[yellow]HTML report limited to {max_html_rows:,} rows "
                f"(file has {len(df):,} differences)[/yellow]"
            )
            df = df.head(max_html_rows)
            is_truncated = True

        # Configure SearchPanes for column filtering
        key_column = df.columns[0]  # First column is always the key
        search_panes_config = SearchPanesConfig.from_settings(
            self.settings,
            list(df.columns),
            key_column
        )

        # Get CDN links
        cdn_links = search_panes_config.get_cdn_links()

        # Generate table rows
        table_rows = ""
        for row in df.iter_rows(named=True):
            row_html = "<tr>"
            for value in row.values():
                # Sanitize HTML
                str_value = str(value) if value is not None else ""
                str_value = str_value.replace("<", "&lt;").replace(">", "&gt;")
                # Convert newlines to <br> for better HTML rendering
                str_value = str_value.replace("\n", "<br>")
                row_html += f"<td>{str_value}</td>"
            row_html += "</tr>"
            table_rows += row_html

        # Generate table headers
        table_headers = ""
        for col in df.columns:
            table_headers += f"<th>{col}</th>"

        # Generate stats
        stats_html = ""
        if summary_stats:
            stats_html = f"""
            <div class="stats">
                <div class="stat-box">
                    <div class="stat-number">{summary_stats.get('total_differences', 0):,}</div>
                    <div class="stat-label">Total Differences</div>
                </div>
                <div class="stat-box">
                    <div class="stat-number">{summary_stats.get('unique_keys', 0):,}</div>
                    <div class="stat-label">Unique Records</div>
                </div>
                <div class="stat-box">
                    <div class="stat-number">{summary_stats.get('exact_matches', 0):,}</div>
                    <div class="stat-label">Exact Matches (Excluded)</div>
                </div>
            </div>
            """

        truncation_warning = ""
        if is_truncated:
            truncation_warning = f"""
            <div class="warning-box">
                Note: This report shows only the first {max_html_rows:,} differences.
                See CSV/Excel files for complete results.
            </div>
            """

        # Build CDN links HTML
        css_links = '<link rel="stylesheet" href="https://cdn.datatables.net/1.13.6/css/jquery.dataTables.min.css">'
        for css_url in cdn_links.get("css", []):
            css_links += f'\n    <link rel="stylesheet" href="{css_url}">'

        js_scripts = '<script src="https://code.jquery.com/jquery-3.7.0.min.js"></script>\n    <script src="https://cdn.datatables.net/1.13.6/js/jquery.dataTables.min.js"></script>'
        for js_url in cdn_links.get("js", []):
            js_scripts += f'\n    <script src="{js_url}"></script>'

        html = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>File Comparison Report</title>
    {css_links}
    {js_scripts}
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 0;
            padding: 20px;
            background: #f5f5f5;
        }}
        .container {{
            max-width: 1600px;
            margin: 0 auto;
            background: white;
            padding: 30px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: #333;
            margin-bottom: 10px;
        }}
        .info {{
            color: #666;
            margin-bottom: 20px;
            font-size: 14px;
        }}
        .stats {{
            display: flex;
            gap: 20px;
            margin-bottom: 30px;
        }}
        .stat-box {{
            flex: 1;
            padding: 20px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border-radius: 8px;
            text-align: center;
        }}
        .stat-number {{
            font-size: 36px;
            font-weight: bold;
            margin-bottom: 5px;
        }}
        .stat-label {{
            font-size: 14px;
            opacity: 0.9;
        }}
        .warning-box {{
            padding: 15px;
            background: #fff3cd;
            border: 1px solid #ffc107;
            border-radius: 4px;
            margin-bottom: 20px;
            color: #856404;
        }}
        table.dataTable {{
            width: 100% !important;
        }}
        table.dataTable thead th {{
            background: #2c3e50;
            color: white;
            padding: 12px;
            text-align: left;
            font-weight: 600;
        }}
        table.dataTable tbody td {{
            padding: 10px;
            border-bottom: 1px solid #eee;
        }}
        .dataTables_wrapper {{
            padding: 20px 0;
        }}
        .dataTables_filter {{
            margin-bottom: 20px;
        }}
        .dataTables_length {{
            margin-bottom: 20px;
        }}

        /* SearchPanes styling */
        .dtsp-searchPanes {{
            margin-bottom: 20px;
        }}
        .dtsp-searchPane {{
            background: white;
            border: 1px solid #e0e0e0;
            border-radius: 4px;
            box-shadow: 0 1px 3px rgba(0,0,0,0.05);
        }}
        .dtsp-title {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 8px 12px;
            font-weight: 600;
            border-radius: 4px 4px 0 0;
        }}
        .dtsp-topRow {{
            background: #f8f9fa;
            padding: 8px;
            border-bottom: 1px solid #e0e0e0;
        }}
        .dtsp-searchCont input {{
            border: 1px solid #ced4da;
            border-radius: 4px;
            padding: 8px 12px;
            width: 100%;
            font-size: 13px;
            transition: all 0.2s ease;
            background: white;
        }}
        .dtsp-searchCont input:focus {{
            outline: none;
            border-color: #667eea;
            box-shadow: 0 0 0 3px rgba(102, 126, 234, 0.1);
        }}
        .dtsp-searchCont input::placeholder {{
            color: #a0a0a0;
            font-style: italic;
            opacity: 0.8;
        }}
        .dtsp-selected {{
            background: #e3f2fd !important;
            border-left: 3px solid #667eea !important;
        }}

        /* Filters Active Counter */
        .dtsp-panesContainer {{
            margin-bottom: 25px;
        }}
        div.dtsp-panesContainer div.dtsp-title {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 10px 16px;
            font-size: 15px;
            font-weight: 600;
            border-radius: 6px;
            box-shadow: 0 2px 4px rgba(102, 126, 234, 0.2);
            margin-bottom: 15px;
            display: inline-block;
            min-width: 180px;
            text-align: center;
        }}
        div.dtsp-panesContainer button.dtsp-clearAll {{
            background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
            color: white;
            border: none;
            padding: 8px 16px;
            border-radius: 4px;
            font-weight: 500;
            cursor: pointer;
            transition: all 0.2s ease;
            box-shadow: 0 2px 4px rgba(245, 87, 108, 0.2);
        }}
        div.dtsp-panesContainer button.dtsp-clearAll:hover {{
            transform: translateY(-1px);
            box-shadow: 0 4px 8px rgba(245, 87, 108, 0.3);
        }}
        div.dtsp-panesContainer button.dtsp-clearAll:active {{
            transform: translateY(0);
        }}
        .help-text {{
            background: #f0f7ff;
            border-left: 4px solid #667eea;
            padding: 12px 15px;
            margin-bottom: 20px;
            color: #1a1a1a;
            border-radius: 4px;
        }}

        /* Pagination styling - add moderate spacing to prevent wrapping */
        .dataTables_paginate {{
            padding-right: 15px !important;
            padding-left: 10px !important;
            margin-top: 20px !important;
        }}
        .dataTables_paginate .paginate_button {{
            margin: 0 5px !important;
            padding: 8px 12px !important;
        }}
        .dataTables_paginate .paginate_button.current {{
            margin: 0 5px !important;
        }}
        .dataTables_paginate .paginate_button.previous,
        .dataTables_paginate .paginate_button.next {{
            margin: 0 8px !important;
        }}
        .dataTables_info {{
            padding-left: 10px !important;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>File Comparison Report</h1>
        <div class="info">
            <strong>Source File:</strong> {source_file}<br>
            <strong>Compared to:</strong> {comparison_file}<br>
            <strong>Generated at:</strong> {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
        </div>

        {stats_html}

        <div class="help-text">
            <strong>💡 Interactive Features:</strong><br>
            • <strong>Column Filters:</strong> {"Use the filter panels above to narrow down results by specific values" if search_panes_config.enabled else "Enable with --enable-search-panes flag"}<br>
            • <strong>Multi-Column Sort:</strong> Hold Shift and click column headers to sort by multiple columns<br>
            • <strong>Search:</strong> Use the search box to find specific text across all columns
        </div>

        {truncation_warning}

        <table id="diffTable" class="display">
            <thead>
                <tr>{table_headers}</tr>
            </thead>
            <tbody>
                {table_rows}
            </tbody>
        </table>
    </div>

    <script>
        $(document).ready(function() {{
            var panePaginationState = {{}};
            var paneScrollState = {{}};
            var isRestoring = false;

            var table = $('#diffTable').DataTable({{
                pageLength: 50,
                order: [[0, 'asc'], [1, 'asc']],
                responsive: true,
                dom: '{search_panes_config.get_dom_layout()}',{search_panes_config.to_js_config()}
            }});

            // Save state frequently (but not during restore)
            setInterval(function() {{
                if (!isRestoring) {{
                    savePaneStates();
                }}
            }}, 50);

            // Restore state after table redraws
            table.on('draw.dt', function() {{
                isRestoring = true;
                setTimeout(function() {{
                    restorePaneStates();
                    setTimeout(function() {{
                        isRestoring = false;
                    }}, 200);
                }}, 100);
            }});

            // Track Clear All and X buttons - only allow reset for these
            $(document).on('click', '.dtsp-clearAll', function() {{
                panePaginationState = {{}};
                paneScrollState = {{}};
            }});

            $(document).on('click', '.clearButton', function() {{
                var $pane = $(this).closest('.dtsp-searchPane');
                var paneId = $pane.find('.dtsp-paneInputButton').attr('placeholder');
                if (paneId) {{
                    delete panePaginationState[paneId];
                    delete paneScrollState[paneId];
                }}
            }});

            // Initial setup
            setTimeout(function() {{
                savePaneStates();
            }}, 200);

            function savePaneStates() {{
                $('.dtsp-searchPane').each(function() {{
                    var $pane = $(this);
                    var paneId = $pane.find('.dtsp-paneInputButton').attr('placeholder');
                    if (!paneId) return;

                    // Save current page number
                    var $paneTable = $pane.find('table');
                    if ($.fn.DataTable.isDataTable($paneTable)) {{
                        try {{
                            var paneTableApi = $paneTable.DataTable();
                            var currentPage = paneTableApi.page();
                            if (currentPage !== undefined && currentPage >= 0) {{
                                panePaginationState[paneId] = currentPage;
                            }}
                        }} catch (e) {{}}
                    }}

                    // Save scroll position
                    var $scrollBody = $pane.find('.dataTables_scrollBody');
                    if ($scrollBody.length) {{
                        var scrollTop = $scrollBody.scrollTop();
                        if (scrollTop >= 0) {{
                            paneScrollState[paneId] = scrollTop;
                        }}
                    }}
                }});
            }}

            function restorePaneStates() {{
                $('.dtsp-searchPane').each(function() {{
                    var $pane = $(this);
                    var paneId = $pane.find('.dtsp-paneInputButton').attr('placeholder');
                    if (!paneId) return;

                    // Always restore to saved page (never reset except Clear All)
                    if (panePaginationState[paneId] !== undefined) {{
                        var $paneTable = $pane.find('table');
                        if ($.fn.DataTable.isDataTable($paneTable)) {{
                            try {{
                                var paneTableApi = $paneTable.DataTable();
                                var savedPage = panePaginationState[paneId];
                                if (paneTableApi.page() !== savedPage) {{
                                    paneTableApi.page(savedPage).draw(false);
                                }}
                            }} catch (e) {{}}
                        }}
                    }}

                    // Always restore scroll position
                    if (paneScrollState[paneId] !== undefined) {{
                        var $scrollBody = $pane.find('.dataTables_scrollBody');
                        if ($scrollBody.length) {{
                            $scrollBody.scrollTop(paneScrollState[paneId]);
                        }}
                    }}
                }});
            }}
        }});
    </script>
</body>
</html>
        """

        output_file.write_text(html, encoding='utf-8')
        console.print(f"[green]HTML report saved:[/green] {output_file}")
        return output_file
